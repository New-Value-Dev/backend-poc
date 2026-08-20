"""XLSX(엑셀) 파서"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.parsing.base import BODY_LEVEL, UnifiedDocument, UnifiedSection


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def sheet_table_text(ws: Worksheet) -> str:
    """빈 행은 건너뛰고, 시트를 다른 파서의 table 섹션과 같은 ` | `/줄바꿈 텍스트로 만든다."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        rows.append(" | ".join(_cell_text(c) for c in row))
    return "\n".join(rows)


class XlsxParser:
    name = "xlsx"
    extensions = ("xlsx",)

    def parse(self, path: Path) -> UnifiedDocument:
        # data_only: 수식이 아니라 마지막에 저장된 계산값을 읽는다 (엑셀에서 열어 저장한 적 없는
        # 파일은 캐시된 값이 없어 수식 문자열이 그대로 나올 수 있음 — openpyxl 자체 한계).
        # read_only: 대용량 시트 메모리 절약.
        wb = load_workbook(str(path), data_only=True, read_only=True)
        try:
            sections: list[UnifiedSection] = []
            order = 0
            sheet_count = len(wb.worksheets)

            for ws in wb.worksheets:
                title = (ws.title or "").strip()
                if title:
                    sections.append(
                        UnifiedSection(
                            content=title, title=title, order=order,
                            level=1, section_type="heading", meta={"sheet": ws.title},
                        )
                    )
                    order += 1

                text = sheet_table_text(ws)
                if text.strip():
                    sections.append(
                        UnifiedSection(
                            content=text, order=order,
                            level=BODY_LEVEL, section_type="table", meta={"sheet": ws.title},
                        )
                    )
                    order += 1
        finally:
            wb.close()

        return UnifiedDocument(
            sections=sections, parser_name=self.name,
            meta={"sheet_count": sheet_count},
        )
